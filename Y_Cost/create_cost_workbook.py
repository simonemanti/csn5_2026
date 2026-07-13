from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path(__file__).with_name("Costi_CSN5_2026.xlsx")

navy = "17365D"
blue = "4472C4"
light_blue = "D9EAF7"
light_green = "E2F0D9"
light_yellow = "FFF2CC"
light_red = "F4CCCC"
white = "FFFFFF"
grey = "E7E6E6"
thin = Side(style="thin", color="B7B7B7")


def title(ws, text, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, text)
    c.fill = PatternFill("solid", fgColor=navy)
    c.font = Font(color=white, bold=True, size=15)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26


def header(ws, row, labels):
    for col, value in enumerate(labels, 1):
        c = ws.cell(row, col, value)
        c.fill = PatternFill("solid", fgColor=blue)
        c.font = Font(color=white, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def widths(ws, values):
    for col, width in enumerate(values, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


wb = Workbook()
ws = wb.active
ws.title = "Dettaglio costi"
title(ws, "CSN5 Grant Giovani 2026 — Dettaglio dei costi", 13)
ws["A2"] = "Compilare le celle gialle. Gli importi sono IVA inclusa; indicare eventuali eccezioni nelle note. Limite: 75.000 € per anno."
ws.merge_cells("A2:M2")
ws["A2"].fill = PatternFill("solid", fgColor=light_yellow)
ws["A2"].alignment = Alignment(wrap_text=True)

labels = ["ID", "Anno", "Categoria", "Componente / attività", "Specifiche o calcolo", "WP", "Costo unitario (€)", "Quantità", "Costo richiesto (€)", "In-kind (€)", "Fornitore / fonte", "Preventivo", "Note / giustificazione"]
header(ws, 4, labels)

rows = [
    ["C01", 1, "Attrezzature", "Cristallo/i analizzatore/i", "Materiale, geometria, dimensioni e intervallo energetico da definire dopo WP1", "WP2", None, None, "=IF(OR(G5=\"\",H5=\"\"),\"\",G5*H5)", 0, "", "Q01", "Ottica selezionata in base al compromesso efficienza–risoluzione"],
    ["C02", 1, "Attrezzature", "Stadi motorizzati e controller", "Numero assi, corsa, carico e ripetibilità", "WP2", None, None, "=IF(OR(G6=\"\",H6=\"\"),\"\",G6*H6)", 0, "", "Q02", "Scansione energetica e allineamento riproducibile"],
    ["C03", 1, "Attrezzature", "Rivelatore SDD ed elettronica", "Area, risoluzione, rate capability e pile-up da giustificare quantitativamente", "WP2", None, 1, "=IF(OR(G7=\"\",H7=\"\"),\"\",G7*H7)", 0, "", "Q03", "Confrontare esplicitamente con gli SDD già disponibili"],
    ["C04", 1, "Attrezzature", "Meccanica e supporti di precisione", "Supporti cristallo, campione e rivelatore; componenti da vuoto se necessari", "WP2", None, None, "=IF(OR(G8=\"\",H8=\"\"),\"\",G8*H8)", 0, "", "Q04", "Integrazione del prototipo"],
    ["C05", 1, "Calcolo", "Workstation di simulazione e analisi", "Configurazione CPU/GPU/RAM motivata dai carichi SHADOW4/Geant4", "WP1", None, 1, "=IF(OR(G9=\"\",H9=\"\"),\"\",G9*H9)", 0, "", "Q05", "Eliminare se sono disponibili risorse adeguate"],
    ["C06", 1, "Consumi", "Campioni e materiali di calibrazione", "Target, standard, finestre, supporti e gas", "WP3", None, 1, "=IF(OR(G10=\"\",H10=\"\"),\"\",G10*H10)", 0, "", "", "Specificare benchmark ed edge >20 keV"],
    ["C07", 1, "Servizi", "Lavorazioni meccaniche specialistiche", "Lavorazioni non realizzabili internamente", "WP2", None, 1, "=IF(OR(G11=\"\",H11=\"\"),\"\",G11*H11)", 0, "", "Q06", "Distinguere dai contributi dei servizi LNF"],
    ["C08", 1, "Missioni", "Campagna sperimentale 1", "Sede; n. persone × n. giorni × costo/giorno + viaggio", "WP3", None, 1, "=IF(OR(G12=\"\",H12=\"\"),\"\",G12*H12)", 0, "", "", "Dettagliare destinazione e scopo"],
    ["C09", 2, "Attrezzature", "Cella operando / ambiente campione", "Geometria, controllo e compatibilità con fluorescenza XAS", "WP4", None, 1, "=IF(OR(G13=\"\",H13=\"\"),\"\",G13*H13)", 0, "", "Q07", "Acquistare solo dopo la validazione della configurazione"],
    ["C10", 2, "Consumi", "Campioni e materiali per misure operando", "Materiali attivi, celle, finestre, elettroliti o gas", "WP4", None, 1, "=IF(OR(G14=\"\",H14=\"\"),\"\",G14*H14)", 0, "", "", "Collegare agli osservabili XAS previsti"],
    ["C11", 2, "Servizi", "Lavorazioni e integrazione finale", "Adeguamenti dopo la validazione del prototipo", "WP2/WP4", None, 1, "=IF(OR(G15=\"\",H15=\"\"),\"\",G15*H15)", 0, "", "Q08", ""],
    ["C12", 2, "Missioni", "Campagne di validazione / operando", "Sede; n. campagne × persone × giorni, viaggio incluso", "WP3/WP4", None, 1, "=IF(OR(G16=\"\",H16=\"\"),\"\",G16*H16)", 0, "", "", "Separare le campagne se hanno sedi o finalità diverse"],
    ["C13", 2, "Missioni", "Conferenza con presentazione orale", "Indicare la conferenza o il criterio di selezione", "Disseminazione", None, 1, "=IF(OR(G17=\"\",H17=\"\"),\"\",G17*H17)", 0, "", "", "Ammissibile se associata a presentazione orale"],
    ["IK01", 1, "In-kind", "Setup VOXES disponibile a LNF", "Tubo X, ottica, schermature, movimentazione e DAQ effettivamente disponibili", "WP2/WP3", 0, 1, 0, None, "INFN-LNF", "Inventario", "Valorizzare solo componenti realmente utilizzabili"],
    ["IK02", 1, "In-kind", "Rivelatori già disponibili", "Modello, area attiva, risoluzione e stato operativo", "WP2/WP3", 0, 1, 0, None, "INFN-LNF", "Inventario", ""],
    ["IK03", 2, "In-kind", "Supporto tecnico e infrastrutture", "Officina, laboratorio, sicurezza e servizi tecnici", "WP2/WP4", 0, 1, 0, None, "INFN-LNF", "Dichiarazione", "Evitare valorizzazioni non documentabili"],
]

for r, values in enumerate(rows, 5):
    for c, value in enumerate(values, 1):
        cell = ws.cell(r, c, value)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for c in (7, 8, 10, 11):
        if ws.cell(r, c).value is None:
            ws.cell(r, c).fill = PatternFill("solid", fgColor=light_yellow)

last = 4 + len(rows)
for r in range(5, last + 1):
    ws.cell(r, 7).number_format = '#,##0.00 [$€-it-IT]'
    ws.cell(r, 9).number_format = '#,##0.00 [$€-it-IT]'
    ws.cell(r, 10).number_format = '#,##0.00 [$€-it-IT]'

dv_year = DataValidation(type="list", formula1='"1,2"')
dv_cat = DataValidation(type="list", formula1='"Attrezzature,Calcolo,Consumi,Servizi,Missioni,Altri costi,In-kind"')
ws.add_data_validation(dv_year)
ws.add_data_validation(dv_cat)
dv_year.add(f"B5:B200")
dv_cat.add(f"C5:C200")
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:M{last}"
widths(ws, [10, 8, 17, 29, 43, 14, 17, 10, 18, 15, 22, 14, 43])

summary = wb.create_sheet("Riepilogo")
title(summary, "Riepilogo finanziario", 5)
summary["A2"] = "Il limite annuo è 75.000 €. Le formule leggono il foglio Dettaglio costi."
summary.merge_cells("A2:E2")
summary["A2"].fill = PatternFill("solid", fgColor=light_yellow)
header(summary, 4, ["Categoria", "Anno 1 (€)", "Anno 2 (€)", "Totale richiesto (€)", "Totale in-kind (€)"])
cats = ["Attrezzature", "Calcolo", "Consumi", "Servizi", "Missioni", "Altri costi"]
for r, cat in enumerate(cats, 5):
    summary.cell(r, 1, cat)
    summary.cell(r, 2, f'=SUMIFS(\'Dettaglio costi\'!$I$5:$I$200,\'Dettaglio costi\'!$C$5:$C$200,$A{r},\'Dettaglio costi\'!$B$5:$B$200,B$4)')
    summary.cell(r, 3, f'=SUMIFS(\'Dettaglio costi\'!$I$5:$I$200,\'Dettaglio costi\'!$C$5:$C$200,$A{r},\'Dettaglio costi\'!$B$5:$B$200,C$4)')
    summary.cell(r, 4, f"=SUM(B{r}:C{r})")
    summary.cell(r, 5, f'=SUMIFS(\'Dettaglio costi\'!$J$5:$J$200,\'Dettaglio costi\'!$C$5:$C$200,$A{r})')
summary["B4"] = "Anno 1"
summary["C4"] = "Anno 2"
total_row = 11
summary.cell(total_row, 1, "TOTALE")
for c in range(2, 6):
    summary.cell(total_row, c, f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}10)")
for c in range(1, 6):
    summary.cell(total_row, c).fill = PatternFill("solid", fgColor=light_green)
    summary.cell(total_row, c).font = Font(bold=True)
    summary.cell(total_row, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
summary["A13"] = "Margine rispetto al limite annuo"
summary["B13"] = "=75000-B11"
summary["C13"] = "=75000-C11"
summary["A14"] = "Verifica limite"
summary["B14"] = '=IF(B11<=75000,"OK","SUPERATO")'
summary["C14"] = '=IF(C11<=75000,"OK","SUPERATO")'
for row in range(5, 15):
    for col in range(1, 6):
        summary.cell(row, col).border = Border(top=thin, bottom=thin, left=thin, right=thin)
for row in range(5, 14):
    for col in range(2, 6):
        summary.cell(row, col).number_format = '#,##0.00 [$€-it-IT]'
summary.conditional_formatting.add("B14:C14", CellIsRule(operator="equal", formula=['"SUPERATO"'], fill=PatternFill("solid", fgColor=light_red)))
widths(summary, [31, 18, 18, 23, 20])
summary.freeze_panes = "A5"

quotes = wb.create_sheet("Preventivi")
title(quotes, "Registro preventivi", 8)
header(quotes, 3, ["Codice", "Voce collegata", "Fornitore", "Data", "Importo IVA incl. (€)", "Validità", "File / riferimento", "Stato e note"])
for r, code in enumerate([f"Q{i:02d}" for i in range(1, 9)], 4):
    quotes.cell(r, 1, code)
    for c in range(1, 9):
        quotes.cell(r, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        quotes.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    for c in range(2, 9):
        quotes.cell(r, c).fill = PatternFill("solid", fgColor=light_yellow)
    quotes.cell(r, 5).number_format = '#,##0.00 [$€-it-IT]'
quotes["A14"] = "Nota"
quotes["B14"] = "Il template richiede un preventivo per le attrezzature e un preventivo dettagliato per ogni voce superiore a 2.000 €."
quotes.merge_cells("B14:H14")
quotes["B14"].alignment = Alignment(wrap_text=True)
quotes["B14"].fill = PatternFill("solid", fgColor=light_yellow)
widths(quotes, [12, 31, 24, 13, 21, 16, 34, 38])
quotes.freeze_panes = "A4"

notes = wb.create_sheet("Istruzioni")
title(notes, "Istruzioni per il budget", 2)
instructions = [
    ("Vincolo", "Massimo 75.000 € per ciascun anno; durata del progetto: 24 mesi."),
    ("Dettaglio", "Riportare la richiesta per ogni anno e descrivere ogni voce in modo verificabile."),
    ("Preventivi", "Allegare un preventivo per le attrezzature e per ogni singola voce superiore a 2.000 €."),
    ("Missioni", "Separare attività sperimentali, riunioni e conferenze. Per le conferenze indicare quella prevista e la presentazione orale."),
    ("In-kind", "Tenere distinti i costi richiesti dalle risorse già disponibili; valorizzare soltanto contributi documentabili."),
    ("Coerenza", "Collegare ogni costo a WP, task, deliverable o requisito tecnico; evitare voci generiche."),
    ("Esclusioni", "Il progetto non può finanziare borse, contratti di ricerca o incarichi a terzi."),
    ("Celle gialle", "Sono campi da compilare o verificare. Le celle con formule non devono essere sovrascritte."),
]
for r, (key, value) in enumerate(instructions, 3):
    notes.cell(r, 1, key).font = Font(bold=True)
    notes.cell(r, 1).fill = PatternFill("solid", fgColor=light_blue)
    notes.cell(r, 2, value)
    for c in (1, 2):
        notes.cell(r, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
        notes.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
widths(notes, [20, 100])

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUT)

# Structural verification after writing.
check = load_workbook(OUT, data_only=False)
assert check.sheetnames == ["Dettaglio costi", "Riepilogo", "Preventivi", "Istruzioni"]
assert check["Riepilogo"]["B11"].value == "=SUM(B5:B10)"
assert check["Dettaglio costi"]["I5"].data_type == "f"
print(OUT)
